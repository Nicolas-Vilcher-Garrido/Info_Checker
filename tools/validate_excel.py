import os
import sys

XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join("exports", "powerbi_export.xlsx")

def main():
    if not os.path.exists(XLSX):
        print(f"[ERRO] Arquivo não encontrado: {os.path.abspath(XLSX)}")
        return 2
    try:
        from openpyxl import load_workbook
    except Exception:
        print("[ERRO] openpyxl não instalado. Rode: pip install openpyxl")
        return 2

    try:
        wb = load_workbook(XLSX, read_only=True, data_only=True)
        print(f"[OK] Abas ({len(wb.sheetnames)}):", ", ".join(wb.sheetnames))
        # tenta acessar a primeira planilha e ler algumas células
        ws = wb[wb.sheetnames[0]]
        rows = 0
        for _ in ws.iter_rows(min_row=1, max_row=10, max_col=10, values_only=True):
            rows += 1
        print(f"[OK] Leituras de teste concluídas ({rows} linhas amostradas).")
        return 0
    except Exception as e:
        print("[ERRO] Falha ao abrir/ler o Excel:", e)
        return 1

if __name__ == "__main__":
    raise SystemExit(main())
